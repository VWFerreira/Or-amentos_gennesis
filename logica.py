from typing import List
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment

BDI = 0.2928
DESCONTO = 0.1013

class Item:
    def __init__(self, item: str, descricao: str, quant: float, unid: str, material: float, mao_obra: float):
        self.item = item
        self.descricao = descricao
        self.quant = quant
        self.unid = unid
        self.material = material
        self.mao_obra = mao_obra
        self.total = (material + mao_obra) * quant * (1 - DESCONTO)
        self.total_com_bdi = self.total * (1 + BDI)

def escrever(ws: Worksheet, row: int, col: int, valor):
    cell = ws.cell(row=row, column=col)
    if col in [5, 6, 7] and isinstance(valor, (int, float)):
        cell.number_format = 'R$ #,##0.00'
    if not any(cell.coordinate in merged for merged in ws.merged_cells.ranges):
        cell.value = valor
    if col == 2:
        cell.alignment = Alignment(horizontal="left", wrap_text=True)

def preencher_planilha(dados: dict) -> BytesIO:
    wb = load_workbook("modelo.xlsx")
    ws = wb.active

    linha_limite = {
        "CIVIL": (10, 57, 57),
        "INSTALAÇÕES ELÉTRICAS": (59, 102, 103),
        "INSTALAÇÕES MECÂNICAS": (105, 121, 121)
    }

    totais = {"material": 0.0, "mao_obra": 0.0, "total_com_bdi": 0.0}

    for tipo, (linha_inicio, linha_fim, linha_subtotal) in linha_limite.items():
        itens = dados.get(tipo, [])
        if not itens:
            continue

        linha_atual = linha_inicio
        subtotal_mat = subtotal_mao = subtotal_total = 0.0

        for item in sorted(itens, key=lambda x: x.item):
            if linha_atual > linha_fim:
                break
            ws.row_dimensions[linha_atual].hidden = False
            escrever(ws, linha_atual, 1, item.item)
            escrever(ws, linha_atual, 2, item.descricao)
            escrever(ws, linha_atual, 3, item.quant)
            escrever(ws, linha_atual, 4, item.unid)
            mat = round(item.material * item.quant, 2)
            mao = round(item.mao_obra * item.quant, 2)
            tot = round(item.total_com_bdi, 2)
            escrever(ws, linha_atual, 5, mat)
            escrever(ws, linha_atual, 6, mao)
            escrever(ws, linha_atual, 7, tot)
            subtotal_mat += mat
            subtotal_mao += mao
            subtotal_total += tot
            linha_atual += 1

        escrever(ws, linha_subtotal, 5, round(subtotal_mat, 2))
        escrever(ws, linha_subtotal, 6, round(subtotal_mao, 2))
        escrever(ws, linha_subtotal, 7, round(subtotal_total, 2))

        totais["material"] += subtotal_mat
        totais["mao_obra"] += subtotal_mao
        totais["total_com_bdi"] += subtotal_total

    escrever(ws, 125, 5, round(totais["material"], 2))
    escrever(ws, 125, 6, round(totais["mao_obra"], 2))
    escrever(ws, 125, 7, round(totais["material"] + totais["mao_obra"], 2))

    escrever(ws, 126, 5, round(totais["material"] * (1 + BDI), 2))
    escrever(ws, 126, 6, round(totais["mao_obra"] * (1 + BDI), 2))
    escrever(ws, 126, 7, round((totais["material"] + totais["mao_obra"]) * (1 + BDI), 2))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def gerar_orcamento_xlsx(itens_selecionados: List[dict], referencia_path: str, nome_arquivo: str = "orcamento_final.xlsx") -> BytesIO:
    df_ref = pd.read_excel(referencia_path)

    dados_por_tipo = {
        "CIVIL": [],
        "INSTALAÇÕES ELÉTRICAS": [],
        "INSTALAÇÕES MECÂNICAS": []
    }

    for entrada in itens_selecionados:
        item_id = entrada["item"]
        quant = entrada["quant"]
        tipo_original = entrada["tipo"].strip().upper()

        if "CIVIL" in tipo_original:
            tipo = "CIVIL"
        elif "ELÉTRICA" in tipo_original:
            tipo = "INSTALAÇÕES ELÉTRICAS"
        elif "MECÂNICA" in tipo_original:
            tipo = "INSTALAÇÕES MECÂNICAS"
        else:
            continue

        linha = df_ref[(df_ref["ITENS"] == item_id) & (df_ref["TIPO"].str.strip().str.upper() == tipo)]
        if linha.empty:
            continue

        linha = linha.iloc[0]
        descricao = linha["DESCRIÇÃO"]
        unid = linha["UNID."]

        material = float(linha["CUSTOS UNITÁRIOS R$MATERIAL"]) if pd.notna(linha["CUSTOS UNITÁRIOS R$MATERIAL"]) else 0
        mao_obra = float(linha["CUSTOS UNITÁRIOS R$MÃO DE OBRA"]) if pd.notna(linha["CUSTOS UNITÁRIOS R$MÃO DE OBRA"]) else 0

        item = Item(item_id, descricao, quant, unid, material, mao_obra)
        dados_por_tipo[tipo].append(item)

    output = preencher_planilha(dados_por_tipo)
    return output
